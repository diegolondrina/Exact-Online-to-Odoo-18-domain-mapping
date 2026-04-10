"""
Apply full formatting to mapping workbooks per output-spec.md.

Colors:
  Direct     = #CFE2F3 (light blue)
  Relational = #D9EAD3 (light green)
  Custom     = #FFF2CC (light yellow)
  Derived    = #FBDAD7 (light red/pink)
  Skip       = #D9D9D9 (grey)
  Header row = #0B5394 (dark blue, white bold text)

Formatting:
  Font: Arial 10pt everywhere
  Header: bold, white, wrap, center h+v, thin bottom border
  Data: wrap, vertical=top, thin bottom border, automatic font color
  Category column: bold in data rows
  Freeze panes: A2
  Auto filter: on header row
  Column widths: standardized
"""

import glob
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Category fills (applied to Category cell only) ---
CATEGORY_FILLS = {
    "direct":     PatternFill(fill_type="solid", fgColor="CFE2F3"),
    "relational": PatternFill(fill_type="solid", fgColor="D9EAD3"),
    "custom":     PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "derived":    PatternFill(fill_type="solid", fgColor="FBDAD7"),
    "skip":       PatternFill(fill_type="solid", fgColor="D9D9D9"),
}

# --- Shared styles ---
HEADER_FILL = PatternFill(fill_type="solid", fgColor="0B5394")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
DATA_FONT = Font(name="Arial", size=10)
DATA_FONT_BOLD = Font(name="Arial", size=10, bold=True)
DATA_ALIGN = Alignment(wrap_text=True, vertical="top")
NO_FILL = PatternFill(fill_type=None)
THIN_BOTTOM = Border(bottom=Side(style="thin", color="FF000000"))

# --- Column widths (A–G) ---
COL_WIDTHS = {
    "A": 25,    # Exact Field
    "B": 14,    # Exact Type
    "C": 12.5,  # Category
    "D": 30,    # Odoo Field
    "E": 12.5,  # Odoo Type
    "F": 26,    # Odoo Model
    "G": 55,    # Notes
}

MAPPINGS_DIR = "mappings"


def format_workbook(path):
    wb = openpyxl.load_workbook(path)

    for ws in wb.worksheets:
        if ws.title.lower() == "legend":
            continue

        header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if header_row is None:
            continue

        # Find Category column
        cat_col = None
        for cell in header_row:
            if cell.value and str(cell.value).strip().lower() == "category":
                cat_col = cell.column
                break

        # --- Header row ---
        for cell in header_row:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BOTTOM

        # --- Data rows ---
        rows_colored = 0
        for row in ws.iter_rows(min_row=2):
            cat_cell = next((c for c in row if c.column == cat_col), None) if cat_col else None

            for cell in row:
                is_cat = cat_cell is not None and cell.column == cat_col
                cell.font = DATA_FONT_BOLD if is_cat else DATA_FONT
                cell.alignment = DATA_ALIGN
                cell.border = THIN_BOTTOM
                cell.fill = NO_FILL

            # Category fill
            if cat_cell and cat_cell.value:
                fill = CATEGORY_FILLS.get(str(cat_cell.value).strip().lower())
                if fill:
                    cat_cell.fill = fill
                    rows_colored += 1

        # --- Column widths ---
        for letter, width in COL_WIDTHS.items():
            ws.column_dimensions[letter].width = width

        # --- Freeze header row ---
        ws.freeze_panes = "A2"

        # --- Auto filter ---
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}1"

        sheet_label = ws.title.encode("ascii", "replace").decode()
        print(f"  [{sheet_label}] - {rows_colored} category cells colored, {ws.max_row - 1} data rows formatted")

    wb.save(path)
    print(f"  Saved: {path}\n")


def main():
    files = [f for f in glob.glob(f"{MAPPINGS_DIR}/*.xlsx") if "example" not in f.lower()]
    if not files:
        print("No .xlsx files found in mappings/")
        return

    print(f"Found {len(files)} workbook(s):\n")
    for path in sorted(files):
        print(f"Processing: {path}")
        format_workbook(path)

    print("Done.")


if __name__ == "__main__":
    main()
