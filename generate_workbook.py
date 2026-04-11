"""
Generate a mapping workbook from CSV source files.

Usage:
    python generate_workbook.py <domain>        # one domain
    python generate_workbook.py --all           # all domains

CSV convention:
    mappings/data/<domain>/ExactEntity-OdooModel.csv
    Filename becomes the sheet title with " → " replacing the first "-":
        CRMAccounts-res.partner.csv  →  sheet "CRMAccounts → res.partner"

    Sheet order is alphabetical by filename.

Each CSV must have the standard 8-column header:
    Exact Field, Exact Type, Category, Odoo Field, Odoo Type, Odoo Model, Related Model, Notes
"""

import argparse
import csv
import io
import os
import sys
import re

# Handle Windows console encoding for arrow characters
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

MAPPINGS_DIR = "mappings"
DATA_DIR = os.path.join(MAPPINGS_DIR, "data")
METADATA_DIR = os.path.join(os.path.dirname(__file__), "metadata")

HEADERS = [
    "Exact Field", "Exact Type", "Category",
    "Odoo Field", "Odoo Type", "Odoo Model", "Related Model", "Notes"
]

LEGEND_ROWS = [
    ("Direct",     "Clear semantic equivalence, at most a type conversion."),
    ("Relational", "Maps to Odoo relational field; requires lookup/resolution."),
    ("Custom",     "No native Odoo equivalent; custom x_aa_ field created."),
    ("Derived",    "No Exact source; computed, defaulted, or inferred."),
    ("Skip",       "Not migrated (denormalized, obsolete, or system-managed)."),
]

# Pattern: ExactEntity-OdooModel.csv (no numeric prefix required).
# ExactEntity is matched non-greedily up to the first hyphen; the rest
# (which may contain dots, e.g. "purchase.order.line") is the Odoo model.
SHEET_NAME_RE = re.compile(
    r"^([^-]+)-(.+)\.csv$"
)


def validate_mapping_filename(filename, domain):
    """Validate that mapping filename follows correct naming convention."""
    m = SHEET_NAME_RE.match(filename)
    if not m:
        sys.exit(f"ERROR: filename '{filename}' does not match ExactEntity-OdooModel.csv convention.")
    
    exact_entity = m.group(1)
    odoo_model = m.group(2)
    
    # Check if Exact entity metadata exists
    exact_metadata_path = os.path.join(METADATA_DIR, "exact", f"{exact_entity}.csv")
    if not os.path.isfile(exact_metadata_path):
        print(f"WARNING: Exact entity metadata file not found: {exact_metadata_path}")
        print("This mapping may be using an entity that hasn't been scraped yet.")
    
    # Check if Odoo model metadata exists
    odoo_metadata_path = os.path.join(METADATA_DIR, "odoo", f"{odoo_model}.csv")
    if not os.path.isfile(odoo_metadata_path):
        sys.exit(f"ERROR: Odoo model metadata file not found: {odoo_metadata_path}")
    
    return exact_entity, odoo_model


def csv_to_sheet_name(filename):
    """Convert CSV filename to sheet title.  e.g. CRMAccounts-res.partner.csv → CRMAccounts → res.partner"""
    m = SHEET_NAME_RE.match(filename)
    if not m:
        sys.exit(f"ERROR: filename '{filename}' does not match ExactEntity-OdooModel.csv convention.")
    exact_entity = m.group(1)
    odoo_model = m.group(2)
    # Truncate long names to fit Excel's 31-character limit
    full_title = f"{exact_entity} → {odoo_model}"
    if len(full_title) > 31:
        # Try to abbreviate the Exact entity name first
        if len(exact_entity) > 15:
            exact_entity = exact_entity[:12] + "..."
        # Then abbreviate the Odoo model if still too long
        full_title = f"{exact_entity} → {odoo_model}"
        if len(full_title) > 31 and len(odoo_model) > 10:
            odoo_model = odoo_model[:7] + "..."
        full_title = f"{exact_entity} → {odoo_model}"
    return full_title


def discover_csvs(domain_dir):
    """Return sorted list of CSV filenames in a domain directory."""
    files = [f for f in os.listdir(domain_dir) if f.endswith(".csv")]
    return sorted(files)


def read_csv(path):
    """Read a mapping CSV and return (headers, rows)."""
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader)
        rows = [row for row in reader if any(cell.strip() for cell in row)]
    return headers, rows


def generate_workbook(domain):
    """Generate a single domain workbook from its CSV files."""
    domain_dir = os.path.join(DATA_DIR, domain)
    if not os.path.isdir(domain_dir):
        sys.exit(f"ERROR: domain directory not found: {domain_dir}")

    csv_files = discover_csvs(domain_dir)
    if not csv_files:
        sys.exit(f"ERROR: no CSV files found in {domain_dir}")

    wb = openpyxl.Workbook()
    first_sheet = True

    for csv_file in csv_files:
        # Validate filename follows correct naming convention
        exact_entity, odoo_model = validate_mapping_filename(csv_file, domain)
        
        sheet_name = csv_to_sheet_name(csv_file)
        csv_path = os.path.join(domain_dir, csv_file)
        headers, rows = read_csv(csv_path)

        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        ws.append(headers)
        for row in rows:
            ws.append(row)

        print(f"  {sheet_name}: {len(rows)} rows")

    # Legend sheet
    lg = wb.create_sheet("Legend")
    lg.append(["Category", "Meaning"])
    for row in LEGEND_ROWS:
        lg.append(list(row))

    output_path = os.path.join(MAPPINGS_DIR, f"{domain}_mapping.xlsx")
    wb.save(output_path)
    print(f"  Saved: {output_path}\n")
    return output_path


def list_domains():
    """Return all domain folder names under DATA_DIR."""
    if not os.path.isdir(DATA_DIR):
        return []
    return sorted(
        d for d in os.listdir(DATA_DIR)
        if os.path.isdir(os.path.join(DATA_DIR, d))
    )


def main():
    parser = argparse.ArgumentParser(description="Generate mapping workbooks from CSV source files.")
    parser.add_argument("domain", nargs="?", help="Domain folder name (e.g. 'subscriptions')")
    parser.add_argument("--all", action="store_true", help="Generate all domains")
    parser.add_argument("--list", action="store_true", help="List available domains")
    args = parser.parse_args()

    if args.list:
        domains = list_domains()
        if domains:
            print("Available domains:")
            for d in domains:
                print(f"  {d}")
        else:
            print(f"No domain folders found in {DATA_DIR}/")
        return

    if args.all:
        domains = list_domains()
        if not domains:
            sys.exit(f"No domain folders found in {DATA_DIR}/")
        print(f"Generating {len(domains)} workbook(s):\n")
        for domain in domains:
            print(f"[{domain}]")
            generate_workbook(domain)
    elif args.domain:
        print(f"[{args.domain}]")
        generate_workbook(args.domain)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
